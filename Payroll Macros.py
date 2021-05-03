import openpyxl
import re
from copy import copy
from datetime import datetime as dt, timedelta
from itertools import cycle

#enter file name
FileName = 'Input.xlsx'
ProcessedName = 'Logs-Dany'
PossibleLocations = {"GA", "WA", "CA", "NJ", "TX", "END"}

COLUMN_WIDTH = 7.5
ROW_HEIGHT = 25

#Color Styles#
header_silver = openpyxl.styles.PatternFill('solid', fgColor="E7F2FF")
corner_black = openpyxl.styles.PatternFill('solid', fgColor="000000")
box_gray = openpyxl.styles.PatternFill('solid', fgColor='EBEBEB')

delta_blue = openpyxl.styles.Font(b=True, i=True, color='0070C0')


no_borders = openpyxl.styles.Border(left=openpyxl.styles.Side(style=None), 
                     right=openpyxl.styles.Side(style=None), 
                     top=openpyxl.styles.Side(style=None), 
                     bottom=openpyxl.styles.Side(style=None))
                     
thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), 
                     right=openpyxl.styles.Side(style='thin'), 
                     top=openpyxl.styles.Side(style='thin'), 
                     bottom=openpyxl.styles.Side(style='thin'))

yellow_ot = openpyxl.styles.PatternFill('solid', fgColor='ffff00')
yellow = openpyxl.styles.Font(b=True, color='ffff00')

oj_sunday = openpyxl.styles.PatternFill('solid', fgColor='ffab40')
oj = openpyxl.styles.Font(b=True, color='ffab40')

red_err = openpyxl.styles.Font(b=True, color="ff0000")
crim_note = openpyxl.styles.Font(b=True, i=True, color="C00000")

bold_delta = openpyxl.styles.Font(b=True, color="000000")

cyan_deduction = openpyxl.styles.PatternFill('solid', fgColor='00ffff')
cyan = openpyxl.styles.Font(b=True, color='00ffff')

purple_timeoff = openpyxl.styles.PatternFill('solid', fgColor='F08BFF')
purple = openpyxl.styles.Font(b=True, color='F08BFF')

no_fill = openpyxl.styles.PatternFill(None, fgColor='ffffff')
no_font = openpyxl.styles.Font(b=True, color="000000")

week_order = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
months = {
    'Jan': '01', 'Feb': '02', 'Mar': '03', 
    'Apr': '04', 'May': '05', 'Jun': '06',  
    'Jul': '07', 'Aug': '08', 'Sep': '09', 
    'Oct': '10', 'Nov': '11', 'Dec': '12'
    }
warehouse_schedule = {
    "WA" : [(8,0),(8,0),(8,0),(16,30),(15,30),(12,0)],
    "CA" : [(8,0),(8,0),(8,0),(16,00),(16,00),(13,0)],
    "GA" : [(8,0),(8,0),(8,0),(15,30),(15,30),(13,0)],
    "NJ" : [(8,0),(8,0),(8,0),(16,30),(15,30),(12,0)]
}

beg_markers = []
dims = {}
extras = ['EX' + str(i) for i in range(1,10)] #Generates 10 extra nom's to be assigned, should be enough.
m = None

# schedule_table = {  #(No. , Name)   (Mon Start, Tue-Fri Start, Sat Start, Mon End, Tue-Fri end, Sat End)
#                    [18,"hongwei"]:[(7,30),(7,30),(7,30),(15,0),(15,0),(12,30)], #GA 7:30-15:00 M-F 	7:30-12:30 SAT
#                    [20,"peterson"]:[(7,30),(7,30),(7,30),(15,0),(15,0),(12,30)] #GA 7:30-15:00 M-F 	7:30-12:30 SAT
#                     }

location = input("Which warehouse location is this? (i.e. 'GA'): ").upper()
while len(location) != 2 or (location not in PossibleLocations):
    location = input("Which warehouse location is this? (i.e. 'GA'): ").upper()
#scheduled time in timedelta; t2f = Tuesday to Friday
mon_b_t = timedelta(hours=warehouse_schedule[location][0][0],minutes=warehouse_schedule[location][0][1])
t2f_b_t = timedelta(hours=warehouse_schedule[location][1][0],minutes=warehouse_schedule[location][1][1])
sat_b_t = timedelta(hours=warehouse_schedule[location][2][0],minutes=warehouse_schedule[location][2][1])
mon_e_t = timedelta(hours=warehouse_schedule[location][3][0],minutes=warehouse_schedule[location][3][1])
t2f_e_t = timedelta(hours=warehouse_schedule[location][4][0],minutes=warehouse_schedule[location][4][1])
sat_e_t = timedelta(hours=warehouse_schedule[location][5][0],minutes=warehouse_schedule[location][5][1])
#used for rounding; set to 5 minutes
res = timedelta(minutes=5)

# while True:
# file_name = input("Enter the filename to be processed: ")
# if file_name is not None:
#     break
# try:


    

wb = openpyxl.load_workbook(FileName)

def get_t_table(location):

    t_table = dict()
    wb3 = openpyxl.load_workbook("Master Schedule.xlsx")
    sheet3 = wb3.active
    for r in range(6,sheet3.max_row + 1):                             #Binary Search implementation?
        if sheet3.cell(row=r,column=1).value == location:
            start_node_r = r
            break
    if start_node_r:
        for r in range(start_node_r + 1, sheet3.max_row + 1):
            if sheet3.cell(row=r,column=1).value in PossibleLocations:
                end_node_r = r
                break

    else:
        raise ValueError("Something went wrong. Checkpoint: Banana")

    for r in range(start_node_r+1, end_node_r):
        if sheet3.cell(row=r, column=1).value:
            k = str(sheet3.cell(row=r, column=1).value)
        else:
            continue

        n = sheet3.cell(row=r, column=3).value

        mbt,met = [dt.strptime(t, "%H:%M") for t in sheet3.cell(row=r, column=4).value.split('-')]
        tbt,tet = [dt.strptime(t, "%H:%M") for t in sheet3.cell(row=r, column=6).value.split('-')]
        sabt,saet = [dt.strptime(t, "%H:%M") for t in sheet3.cell(row=r, column=9).value.split('-')]

        s1,e1 = timedelta(hours=mbt.hour, minutes=mbt.minute), timedelta(hours=met.hour, minutes=met.minute)
        s2,e2 = timedelta(hours=tbt.hour, minutes=tbt.minute), timedelta(hours=tet.hour, minutes=tet.minute)
        s3,e3 = timedelta(hours=sabt.hour, minutes=sabt.minute), timedelta(hours=saet.hour, minutes=saet.minute)

        t_table[k] = [n, s1, e1, s2, e2, s3, e3]
    return t_table                                   # {no. : [Name, Start1,end1, start2, end2, start3, end3 ]}
    

def strfdelta(tdelta, fmt):
    d = {"days": tdelta.days}
    d["hours"], rem = divmod(tdelta.seconds, 3600)
    d["minutes"], d["seconds"] = divmod(rem, 60)
    return fmt.format(**d)


def smart_select(src):
    index_lst = [ws.lower()[:5] for ws in src.sheetnames]
    if 'logs-' in index_lst:
        pos = index_lst.index('logs-')
    elif 'logs ' in index_lst:
        pos = index_lst.index('logs ')
    else:
        print("No revised sheet detected; using default Logs.")
        return 'Logs'
    print(f"Input sheet selected: {src.sheetnames[pos]}")
    return src.sheetnames[pos]

a = smart_select(wb)
source = wb[a]
sheet = wb.copy_worksheet(source)
sheet.title = ProcessedName

calendar = dict.fromkeys(range(1,32), None)
# print(wb.sheetnames)
# print(sheet['A1'].value)
# print(sheet.cell(row=5, column=1).value == 'No :')
def regex_clean(s):
    #ensures that all pinyin symbols specifically the colons are transformed into English ones.
    #does the additional step of stripping any space that is added during handling process

    return re.sub(r'[^\w\s]',':',str(s))



def set_border(ws, origin, end, double=False):
    rows = ws[f"{origin}:{end}"]
    if double:
        side_style = 'double'
    else:
        side_style = 'thin'
    for row in rows:
        row[0].border = openpyxl.styles.Border(left=openpyxl.styles.Side(style=side_style))
        row[-1].border = openpyxl.styles.Border(right=openpyxl.styles.Side(style=side_style))
    for c in rows[0]:
        if c == rows[0][0]:
            c.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style=side_style),left=openpyxl.styles.Side(style=side_style))
        elif c == rows[0][-1]:
            c.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style=side_style),right=openpyxl.styles.Side(style=side_style))
        else:
            c.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style=side_style))
    for c in rows[-1]:
        if c == rows[-1][0] :
            c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style=side_style),left=openpyxl.styles.Side(style=side_style))
        elif c == rows[-1][-1]:
            c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style=side_style),right=openpyxl.styles.Side(style=side_style))
        else:
            c.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style=side_style))

def header():
    sheet.cell(row=1,column=1).font = no_font
    sheet.cell(row=1,column=1).alignment = openpyxl.styles.Alignment(horizontal='left') 
    try:
        sheet.unmerge_cells('A1:AE2')
    except:
        pass
    sheet.cell(row=1,column=1).value = "Yellow highlights"
    sheet.cell(row=1,column=1).font = yellow
    sheet.cell(row=2,column=1).value = "are assumed OT."
    sheet.cell(row=1,column=4).value = "Cyan highlights"
    sheet.cell(row=1,column=4).font = cyan
    sheet.cell(row=2,column=4).value = "are Saturdays with lunch deducted."
    sheet.cell(row=1,column=9).value = "Orange highlights"
    sheet.cell(row=1,column=9).font = oj
    sheet.cell(row=2,column=9).value = "are hours worked on Sunday. (?)"
    sheet.cell(row=1,column=14).value = "Purple highlights"
    sheet.cell(row=1,column=14).font = purple
    sheet.cell(row=2,column=14).value = "days where employees left early; verify time-off."
    sheet.cell(row=1,column=20).value = "Red highlights"
    sheet.cell(row=1,column=20).font = red_err
    sheet.cell(row=2,column=20).value = "are missing time or information. Need to verify."

def dating(lst,cal=calendar):
    #Depending on user input; will shift the days of the week for cycle zip later.
    first_day = input("Enter the first date of the month's day of the week? (i.e. Monday or Mon): ")
    while (len(first_day) < 3) or (first_day[:3].capitalize() not in week_order):
        first_day = input("Enter the first date's day of the week? (i.e. Monday or Mon): ")
    else:
        while week_order[0] != first_day[:3].capitalize():
            week_order.append(week_order.pop(0))

    #creates a month dictionary with dates as key and days as value.
    cal = {k:v for k,v in zip(cal.keys(),cycle(week_order))}

    dates_r = [sheet[coord].row-1 for coord in lst]
    for row in dates_r:
        for col in range(2, 19):
            cel = sheet.cell(row=row,column=col).value
                        #reformat the date header with dates
            if cel in cal.keys():
                sheet.cell(row=row,column=col).value = f"{cal[cel]} {cel}"
    return cal

#First, we find how many rows of employee datas there are
def beg_markers_calc(lst):
    row_range = sheet[1:sheet.max_row]
    for r in row_range:
        if r[0].value == 'No :': #if first cell of the row is 'No :'
            lst.append(r[0].coordinate)

def sub_header(lst):
    r = [sheet[coord].row for coord in lst]
    for coord in r:
        for i in range(1,21):
            if i == 6:
                continue
            else:
                sheet.cell(row=coord,column=i).fill = header_silver
        sheet.cell(row=coord,column=4).value = "Prev. Hrs:"
        sheet.cell(row=coord,column=4).font = bold_delta
        sheet.cell(row=coord,column=6).border = thin_border
        sheet.cell(row=coord,column=6).fill = box_gray
        sheet.cell(row=coord,column=6).font = delta_blue
        sheet.cell(row=coord,column=6).number_format = "[h]:mm"
        sheet.cell(row=coord,column=19).value = ""
        sheet.cell(row=coord,column=19).font = bold_delta
        sheet.cell(row=coord,column=21).value = ""
        sheet.cell(row=coord,column=21).font = bold_delta
        sheet.cell(row=coord,column=22).font = no_font
        sheet.cell(row=coord,column=22).fill = no_fill
        sheet.cell(row=coord,column=23).value = "Comments (internal use):"
        sheet.cell(row=coord,column=21).font = no_font
        sheet.cell(row=coord,column=6).number_format = "[h]:mm"

#Inserting blank rows; so we can use this for data later
def space_prep(lst):
    for coord in lst[::-1]:
        for _ in range(10): #repeats 10 times; add 10 rows
            sheet.insert_rows(sheet[coord].row+1)
        for _ in range(3): #adds space to above the calendar header
            sheet.insert_rows(sheet[coord].row-1)
    lst.clear() #clears the set, to re-calculate the beginning of rows for data insertion.

def row_headers(lst):
    global m
    sheet.insert_cols(1)
    r = [sheet[coord].row for coord in lst]
    m = input("Enter the month (i.e. April or Apr): ")
    while (len(m) < 3) or (m[:3].capitalize() not in months.keys()):
        m = input("Enter the month (i.e. April or Apr): ")
    m = m[:3].capitalize()
    for coord in r:
    #Adds month to top left square
        sheet.cell(row=coord-1,column=1).value = m.title()
        sheet.cell(row=coord-1,column=1).font = bold_delta
        sheet.cell(row=coord-1,column=1).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row=coord-1,column=1).alignment = openpyxl.styles.Alignment(vertical='center')
        sheet.cell(row=coord-1,column=1).fill = no_fill
    #Adding row headers
        sheet.cell(row=coord+1,column=1).value = "Ck1:"
        sheet.cell(row=coord+2,column=1).value = "Ck2:"
        sheet.cell(row=coord+3,column=1).value = "In:"
        sheet.cell(row=coord+4,column=1).value = "Out:"
        sheet.cell(row=coord+5,column=1).value = "Hrs:"
        sheet.cell(row=coord+6,column=1).value = "Rgr:"
        sheet.cell(row=coord+7,column=1).value = "OT:"
        sheet.cell(row=coord+11,column=1).value = "Raw:"

def delta(b,a,dy=""):
    if dy not in week_order:
        raise TypeError("Need a day of the week in the delta function.")
    #resets variable
    styl = no_fill
    #Parse string and convert to delta -> b2
    b1 = dt.strptime(b,"%H:%M")
    b2 = timedelta(hours=b1.hour, minutes=b1.minute)
    #Parse string and convert to delta -> a2
    a1 = dt.strptime(a,"%H:%M")
    a2 = timedelta(hours=a1.hour, minutes=a1.minute)

    s = (b2-a2).total_seconds()
    if s > 18000: #if work is greater than 5 hours (18000 seconds), we must deduct lunch time.
        s -= 1800
        if dy == "Sat": styl = cyan_deduction
    hh, ss = divmod(s, 3600) #converts raw total seconds of a_delta to hours.
    mm = ss / 60 #converts remainder seconds to minutes; there should be no remainder
    return '{:02}:{:02}'.format(int(hh), int(mm)), styl

def unique_sch(sch,sp,num=None):
    sp_dict = {
        "mb" : 1,
        "tb" : 3,
        "sb" : 5,
        "me" : 2,
        "te" : 4,
        "se" : 6                           
        }
    if t_table.get(num):
        index = sp_dict[sp]
        return t_table[num][index]
    else:
        return sch

def splitter(lst):
    beg_r = [sheet[coord].row+1 for coord in lst]
    end_r = [sheet[coord].row+2 for coord in lst]
    nend_r = [sheet[coord].row+4 for coord in lst]
    nbeg_r = [sheet[coord].row+3 for coord in lst]
    dif_r = [sheet[coord].row+5 for coord in lst]
    tbs_r = [sheet[coord].row+11 for coord in lst]
    day_r = [sheet[coord].row-1 for coord in lst]
    tbs = list(zip(tbs_r, beg_r, end_r, nbeg_r, nend_r, dif_r, day_r))
    for emp_t in tbs:
        for col in range(2, 18):
            cel = regex_clean(sheet.cell(row=emp_t[0],column=col).value)
            if cel is not None:
                tbi = cel.strip().split()

                #converts all working cells to correct format before inputting data
                for num in range(1,5):
                    sheet.cell(row=emp_t[num],column=col).number_format = '[h]:mm'

                if len(tbi) == 2:
                    #original time spliced and set
                    sheet.cell(row=emp_t[1],column=col).value = tbi[0]
                    sheet.cell(row=emp_t[2],column=col).value = tbi[1]

                    #converting time to delta time to insert into rounded time
                    b_t = dt.strptime(tbi[0], "%H:%M")
                    b_delta = timedelta(hours=b_t.hour, minutes=b_t.minute)
                    e_t = dt.strptime(tbi[1], "%H:%M")
                    e_delta = timedelta(hours=e_t.hour, minutes=e_t.minute)

                    #This step is to catch any manually entered end-time that is using 12hr format instead of 24hr. If the end-time is less than beginning clock-in time, then most i.e. 3:00 for 15:00 PM, we will add 12 hours.
                    if e_delta < b_delta:
                        e_delta += timedelta(hours=12)

                    #storing the day of the week to be used later for lunch deduction highlight
                    day = sheet.cell(row=emp_t[6],column=col).value[:3]

                    #take note of the current employee's  No.
                    num = sheet.cell(row=emp_t[1]-1, column=4).value

                    #checking the day of the week; in order to know which time to compare to
                    if day == "Mon":
                        sheet.cell(row=emp_t[3],column=col).value, sheet.cell(row=emp_t[3],column=col).fill = t_round(unique_sch(mon_b_t, 'mb', num), 'beg', b_delta) #returns monday rounded beginning time
                        sheet.cell(row=emp_t[4],column=col).value, sheet.cell(row=emp_t[4],column=col).fill = t_round(unique_sch(mon_e_t, 'me', num), 'end', e_delta)
                    elif day == "Sat":
                        sheet.cell(row=emp_t[3],column=col).value, sheet.cell(row=emp_t[3],column=col).fill = t_round(unique_sch(sat_b_t, 'sb', num), 'beg', b_delta)
                        sheet.cell(row=emp_t[4],column=col).value, sheet.cell(row=emp_t[4],column=col).fill = t_round(unique_sch(sat_e_t, 'se', num), 'end', e_delta)
                    elif (day in week_order) and (day != "Sun"): #assuming that it's Tuesday-Friday
                        sheet.cell(row=emp_t[3],column=col).value, sheet.cell(row=emp_t[3],column=col).fill = t_round(unique_sch(t2f_b_t, 'tb', num), 'beg', b_delta)
                        sheet.cell(row=emp_t[4],column=col).value, sheet.cell(row=emp_t[4],column=col).fill = t_round(unique_sch(t2f_e_t, 'te', num), 'end', e_delta)
                    else:
                        #Sundays are not expected; script will not round and then fill it orange.
                        _ = input("Just so you know! There is a Sunday work hours. Please investigate. If there's Sunday, OT total needs to be recalculated.")
                        sheet.cell(row=emp_t[3],column=col).value, sheet.cell(row=emp_t[3],column=col).fill = f"SUN!{tbi[0]}", oj_sunday
                        sheet.cell(row=emp_t[4],column=col).value, sheet.cell(row=emp_t[4],column=col).fill = f"SUN!{tbi[1]}", oj_sunday
                    
                    #If there's a beginning and end time, also subtract to get the difference.
                    sheet.cell(row=emp_t[5],column=col).number_format = "[h]:mm"      
                    sheet.cell(row=emp_t[5],column=col).value, sheet.cell(row=emp_t[5],column=col).fill = delta(sheet.cell(row=emp_t[4],column=col).value, sheet.cell(row=emp_t[3],column=col).value, day)
                    sheet.cell(row=emp_t[5],column=col).font = bold_delta
                    

                else:
                    #Change values to Error, then red + bold the text.
                    sheet.cell(row=emp_t[1],column=col).value = "ERR"
                    sheet.cell(row=emp_t[2],column=col).value = "ERR"


def t_round(sch_delta,typ,a_delta):
    if typ not in {'beg','end'}:
        raise ValueError("Type must be specified as 'beg' or 'end'.")
    rd_true = 0 #resets rd_true variable; just in case
    styl = no_fill #resets default styling as no fill
    if typ == 'beg': #calcualting beginning hours using beginning rounding method -> up 5
        if a_delta > sch_delta: #if late
            rd_true = a_delta.total_seconds() % res.total_seconds()
            if rd_true:
                a_delta += timedelta(seconds=int(300 - rd_true)) #Adds the amount needed to get to next nearest 5 minutes
        elif a_delta < sch_delta:
            if (sch_delta.total_seconds() - a_delta.total_seconds()) / 60 >= 30: #if earlier than 30 minutes; need to highlight after rounding
                styl = yellow_ot
                rd_true = a_delta.total_seconds() % res.total_seconds()
                if rd_true:
                    a_delta += timedelta(seconds=int(300 - rd_true)) #Adds the amount needed to get to next nearest 5 minutes
                    
            else:
                a_delta = sch_delta

    elif typ == 'end': 
        if a_delta > sch_delta: #if worked longer than scheduled hours
            if (a_delta.total_seconds() - sch_delta.total_seconds()) / 60 >= 15: #15 minutes end or more.
                styl = yellow_ot
                rd_true = a_delta.total_seconds() % res.total_seconds()
                if rd_true:
                    a_delta -= timedelta(seconds=int(rd_true)) #Subtract the modulo to get to the floored nearest 5 minutes.
            else: #if greater; but not 30 minutes or more. We assume there is no over-time.
                a_delta = sch_delta
        elif a_delta < sch_delta: #if less than scheduled time; we assume that they left work early, floor to the last 5 minutes.
            styl = purple_timeoff
            rd_true = a_delta.total_seconds() % res.total_seconds()
            if rd_true:
                a_delta -= timedelta(seconds=int(rd_true))
    s = a_delta.total_seconds()
    hh, ss = divmod(s, 3600) #converts raw total seconds of a_delta to hours.
    mm = ss / 60 #converts remainder seconds to minutes; there should be no remainder
    # print('{:02}:{:02}'.format(int(hh), int(mm)))
    # print('{:02d}:{:02d}'.format(int(hh), int(mm)))
    return '{:02}:{:02}'.format(int(hh), int(mm)), styl

def resize():
    for i in range(2,26):
        c = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[c].width = COLUMN_WIDTH
    for i in range(4,sheet.max_row+1):
        sheet.row_dimensions[i].height = ROW_HEIGHT

def ot_calc(lst):
    r = [sheet[coord].row for coord in lst]
    day = None
    for coord in r:
        sum_nodes = []
        ot_nodes =[]
        d_r = coord - 1
        total_r = coord + 5
        rgr_r = coord + 6
        ot_r = coord + 7
        last_sunday = None
        for col in range(2, 19):
            try:
                day = sheet.cell(row=d_r,column=col).value[:3]
            except TypeError: #If there is no more date, then we have reached the last day of the payroll.
                if (last_sunday is not None) and (col-1 != last_sunday): #If there is a previous Sunday, meaning that the day after we have not calc OT for yet. Same calculation using day after last Sunday and one day before calendar runs out)
                    offset_col_start = last_sunday.column + 1
                    if offset_col_start + 1 != col: #If there is a day between the last Sunday on the payroll and the end date, then we can do calculation"
                        string_formula = sheet.cell(row=total_r,column=offset_col_start).coordinate #sets string formula to the first number; keep in mind .
                        sheet.cell(row=cel.row+1,column=offset_col_start).number_format = '[h]:mm'
                        sheet.cell(row=cel.row+1,column=offset_col_start).value = f'={sheet.cell(row=total_r,column=offset_col_start).coordinate}-{sheet.cell(row=ot_r,column=offset_col_start).coordinate}'
                        sheet.cell(row=cel.row+2,column=offset_col_start).number_format = '[h]:mm'
                        sheet.cell(row=cel.row+2,column=offset_col_start).value = f'=IF({sheet.cell(row=total_r,column=offset_col_start).coordinate}-"08:00">0, {sheet.cell(row=total_r,column=offset_col_start).coordinate}-"08:00", 0)'
                        start_coord = sheet.cell(row=total_r,column=offset_col_start+1).coordinate
                        end_coord = sheet.cell(row=total_r,column=col-1).coordinate
                else:
                    continue #Break out of loop and move on to next col

                if start_coord != end_coord:
                    for tple in sheet[f"{start_coord}:{end_coord}"]:
                        for cel in tple:
                            sheet.cell(row=cel.row+1,column=cel.column).number_format = '[h]:mm'
                            sheet.cell(row=cel.row+1,column=cel.column).value = f'={cel.coordinate}-{sheet.cell(row=cel.row+2,column=cel.column).coordinate}'
                            sheet.cell(row=cel.row+2,column=cel.column).number_format = '[h]:mm'
                            sheet.cell(row=cel.row+2,column=cel.column).value = f'=IF({cel.coordinate}-"08:00">0, {cel.coordinate}-"08:00", 0)'
                            insert = "+" + cel.coordinate
                            string_formula = string_formula + insert
                else:
                    sheet.cell(row=total_r,column=offset_col_start).number_format = '[h]:mm'
                    sheet.cell(row=total_r,column=offset_col_start).value = f'=IF({sheet.cell(row=total_r,column=offset_col_start).coordinate}-"08:00">0, {sheet.cell(row=total_r,column=offset_col_start).coordinate}-"08:00", 0)'
                    string_formula = string_formula + f"+{start_coord}"


                if sheet.cell(row=total_r, column=col).value != string_formula:
                    sheet.cell(row=total_r, column=col).number_format = '[h]:mm'
                    sheet.cell(row=total_r+1, column=col).number_format = '[h]:mm'
                    sheet.cell(row=total_r+2, column=col).number_format = '[h]:mm'
                    sheet.cell(row=total_r, column=col).font = delta_blue
                    sheet.cell(row=total_r, column=col).value = f"={string_formula}"
                    sheet.cell(row=total_r+1, column=col).font = delta_blue
                    sheet.cell(row=total_r+1, column=col).value = f'={sheet.cell(row=total_r, column=col).coordinate}-{sheet.cell(row=ot_r,column=col).coordinate}'
                    sheet.cell(row=total_r+2, column=col).font = delta_blue
                    sheet.cell(row=total_r+2, column=col).value = f'=IF({sheet.cell(row=total_r, column=col).coordinate}-"40:00">0, {sheet.cell(row=total_r, column=col).coordinate}-"40:00", 0)'
                last_sunday = None #End of the calculation, reset last_sunday for next person's payroll.
                sum_nodes.append(sheet.cell(row=total_r, column=col).coordinate)
                ot_nodes.append(sheet.cell(row=ot_r, column=col).coordinate)
        
            # If it's Sunday (beginning to middle back check)
            if day == "Sun": #We only start calculating when its sunday.
                offset_col_start = col - 6
                offset_col_end = col - 1
                while offset_col_start < 2: #bringing up the minimum so theres no out of bound index error
                    offset_col_start += 1
                while offset_col_end < 2: #bringing up the minimum so theres no out of bound index error
                    offset_col_end += 1
                # print(f"cp: {offset_col_end}")
                if offset_col_end != col:
                    string_formula = sheet.cell(row=total_r,column=offset_col_start).coordinate
                    sheet.cell(row=total_r+1,column=offset_col_start).number_format = '[h]:mm'
                    sheet.cell(row=total_r+1,column=offset_col_start).value = f'={sheet.cell(row=total_r,column=offset_col_start).coordinate}-{sheet.cell(row=ot_r,column=offset_col_start).coordinate}'
                    sheet.cell(row=total_r+2,column=offset_col_start).number_format = '[h]:mm'
                    sheet.cell(row=total_r+2,column=offset_col_start).value = f'=IF({sheet.cell(row=total_r,column=offset_col_start).coordinate}-"08:00">0, {sheet.cell(row=total_r,column=offset_col_start).coordinate}-"08:00", 0)'
                    start_coord = sheet.cell(row=total_r,column=offset_col_start+1).coordinate
                    end_coord = sheet.cell(row=total_r,column=offset_col_end).coordinate
                    if start_coord != end_coord and offset_col_start != sheet[end_coord].column:
                        for tple in sheet[f"{start_coord}:{end_coord}"]:
                            for cel in tple:
                                sheet.cell(row=cel.row+1,column=cel.column).number_format = '[h]:mm'
                                sheet.cell(row=cel.row+1,column=cel.column).value = f'={sheet.cell(row=total_r,column=cel.column).coordinate}-{sheet.cell(row=ot_r,column=cel.column).coordinate}'
                                sheet.cell(row=cel.row+2,column=cel.column).number_format = '[h]:mm'
                                sheet.cell(row=cel.row+2,column=cel.column).value = f'=IF({sheet.cell(row=cel.row,column=cel.column).coordinate}-"08:00">0, {sheet.cell(row=cel.row,column=cel.column).coordinate}-"08:00", 0)'
                                insert = "+" + cel.coordinate
                                string_formula = string_formula + insert
                    elif start_coord != end_coord and offset_col_start == sheet[end_coord].column:
                        pass
                    else: #if start_coord == end_coord, aka First day of the payroll is a Saturday, then Sunday calculation is only string_formula
                        sheet.cell(row=total_r+1,column=sheet[end_coord].column).number_format = '[h]:mm'
                        sheet.cell(row=total_r+1,column=sheet[end_coord].column).value = f'={sheet.cell(row=total_r,column=sheet[end_coord].column).coordinate}-{sheet.cell(row=ot_r,column=sheet[end_coord].column).coordinate}'
                        sheet.cell(row=total_r+2,column=sheet[end_coord].column).number_format = '[h]:mm'
                        sheet.cell(row=total_r+2,column=sheet[end_coord].column).value = f'=IF({sheet.cell(row=total_r,column=sheet[end_coord].column).coordinate}-"08:00">0, {sheet.cell(row=total_r,column=sheet[end_coord].column).coordinate}-"08:00", 0)'
                        string_formula = string_formula + f"+ {end_coord}"
                    if sheet.cell(row=total_r, column=col).value != string_formula:
                        sheet.cell(row=total_r, column=col).number_format = '[h]:mm'
                        sheet.cell(row=total_r+1, column=col).number_format = '[h]:mm'
                        sheet.cell(row=total_r+2, column=col).number_format = '[h]:mm'
                        sheet.cell(row=total_r, column=col).font = delta_blue
                        sheet.cell(row=total_r, column=col).value = f"={string_formula}"
                        sheet.cell(row=total_r+1, column=col).font = delta_blue
                        sheet.cell(row=total_r+2, column=col).font = delta_blue
                        if last_sunday is None: #This line adds the overflow, extra partial hrs from last week.
                            sheet.cell(row=ot_r, column=col).value = f'=IF({sheet.cell(row=total_r, column=col).coordinate}+{sheet.cell(row=total_r-5, column=7).coordinate}-"40:00">0, {sheet.cell(row=total_r, column=col).coordinate}+{sheet.cell(row=total_r-5, column=7).coordinate}-"40:00", 0)'
                        else:
                            sheet.cell(row=ot_r, column=col).value = f'=IF({sheet.cell(row=total_r, column=col).coordinate}-"40:00">0, {sheet.cell(row=total_r, column=col).coordinate}-"40:00", 0)'
                        sheet.cell(row=rgr_r, column=col).value = f'={sheet.cell(row=total_r,column=col).coordinate}-{sheet.cell(row=ot_r,column=col).coordinate}'#Regardless of overflow, need to add Rgr row
                        last_sunday = sheet.cell(row=total_r,column=col)
                sum_nodes.append(sheet.cell(row=total_r, column=col).coordinate)
                ot_nodes.append(sheet.cell(row=ot_r, column=col).coordinate)
        if len(sum_nodes) == 1:
            total_formula = f"={sum_nodes[0]}"
            sheet.cell(row=(coord),column=19).value = total_formula
            ot_formula = f"={ot_nodes[0]}"
            sheet.cell(row=(coord),column=17).value = ot_formula #1 OT slot possible
            reg_formula = f"={sheet.cell(row=(coord),column=19).coordinate}-{sheet.cell(row=(coord),column=17).coordinate}"
            sheet.cell(row=(coord),column=15).value = reg_formula
        if len(sum_nodes) == 2:
            total_formula = f"={sum_nodes[0]}+{sum_nodes[1]}"
            sheet.cell(row=(coord),column=19).value = total_formula
            ot_formula = f"={ot_nodes[0]}+{ot_nodes[1]}"
            sheet.cell(row=(coord),column=17).value = ot_formula #2 OT slot possible
            reg_formula = f"={sheet.cell(row=(coord),column=19).coordinate}-{sheet.cell(row=(coord),column=17).coordinate}"
            sheet.cell(row=(coord),column=15).value = reg_formula
        if len(sum_nodes) == 3:
            total_formula = f"={sum_nodes[0]}+{sum_nodes[1]}+{sum_nodes[2]}"
            sheet.cell(row=(coord),column=19).value = total_formula
            ot_formula = f"={ot_nodes[0]}+{ot_nodes[1]}"
            sheet.cell(row=(coord),column=17).value = ot_formula #2 OT slot possible
            reg_formula = f"={sheet.cell(row=(coord),column=19).coordinate}-{sheet.cell(row=(coord),column=17).coordinate}"
            sheet.cell(row=(coord),column=15).value = reg_formula

def post_format(lst):
    r = reversed([sheet[coord].row for coord in lst])
    for coord in r:
        tp_r = coord + 8
        md_r = coord + 9
        bt_r = coord + 10
        signature_r = coord + 3
        date_r = coord + 6

        if sheet.cell(row=coord,column=4).value:
            num = str(sheet.cell(row=coord,column=4).value)
        else:
            num = None
        if t_table.get(num):
            sheet.cell(row=coord+3,column=24).value = 'Mon:'
            sheet.cell(row=coord+3,column=25).value = f'{strfdelta(t_table[num][1], "{hours:02d}:{minutes:02d}")}-{strfdelta(t_table[num][2], "{hours:02d}:{minutes:02d}")}'
            sheet.cell(row=coord+4,column=24).value = 'Tue-Fri:'
            sheet.cell(row=coord+4,column=25).value = f'{strfdelta(t_table[num][3], "{hours:02d}:{minutes:02d}")}-{strfdelta(t_table[num][4], "{hours:02d}:{minutes:02d}")}'
            sheet.cell(row=coord+5,column=24).value = 'Sat:'
            sheet.cell(row=coord+5,column=25).value = f'{strfdelta(t_table[num][5], "{hours:02d}:{minutes:02d}")}-{strfdelta(t_table[num][6], "{hours:02d}:{minutes:02d}")}'

        #Compliance legal agreement, center the text
        sheet.cell(row=tp_r,column=2).value = "I attest that the hours I recorded as my time worked are accurate.I accurately recorded"
        sheet.cell(row=md_r,column=2).value = "all time worked and did not complete any required work duties outside of the recorded time."
        sheet.cell(row=bt_r,column=2).value = "I received all meal and rest periods that I am entitled to during this time period."
        sheet.cell(row=tp_r,column=2).alignment = openpyxl.styles.Alignment(vertical='center')
        sheet.cell(row=md_r,column=2).alignment = openpyxl.styles.Alignment(vertical='center')
        sheet.cell(row=bt_r,column=2).alignment = openpyxl.styles.Alignment(vertical='center')

        #Memo box
        sheet.cell(row=tp_r,column=13).value = "Memo:"
        sheet.cell(row=tp_r,column=13).alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet.cell(row=tp_r,column=13).alignment = openpyxl.styles.Alignment(vertical='center')
        sheet.cell(row=tp_r,column=13).font = bold_delta

        #border function setting for memo box
        set_border(sheet, sheet.cell(row=tp_r,column=13).coordinate, sheet.cell(row=bt_r,column=20).coordinate)

        #signature and date title on top right corner
        sheet.cell(row=signature_r,column=20).value = "Signature:"
        sheet.cell(row=signature_r,column=20).alignment = openpyxl.styles.Alignment(vertical='top')
        sheet.cell(row=date_r,column=20).value = "Date:"
        sheet.cell(row=date_r,column=20).alignment = openpyxl.styles.Alignment(vertical='top')
        sheet.cell(row=signature_r,column=20).font = bold_delta
        sheet.cell(row=date_r,column=20).font = bold_delta

        #Regular, OT and total information on top.
        sheet.cell(row=(coord),column=14).value = "Regular:"
        sheet.cell(row=(coord),column=16).value = "OT:"
        sheet.cell(row=(coord),column=18).value = "Total:"
        sheet.cell(row=(coord),column=14).font = bold_delta
        sheet.cell(row=(coord),column=16).font = bold_delta
        sheet.cell(row=(coord),column=18).font = bold_delta

        #colors the data responding regular OT and total information font bold and red. Also change to [h]:mm format
        sheet.cell(row=(coord),column=15).font = red_err
        sheet.cell(row=(coord),column=17).font = red_err
        sheet.cell(row=(coord),column=19).font = red_err
        sheet.cell(row=(coord),column=15).number_format = '[h]:mm'
        sheet.cell(row=(coord),column=17).number_format = '[h]:mm'
        sheet.cell(row=(coord),column=19).number_format = '[h]:mm'

        #hides the split two rows; raw data can still be looked at for 
        sheet.row_dimensions.group(coord+1, coord+2, hidden=True)
        #Paints blue on the row headers
        for i in range(13):
            sheet.cell(row=(coord+i),column=1).hyperlink = None
            sheet.cell(row=(coord+i),column=1).font, sheet.cell(row=(coord+i),column=1).fill = bold_delta, header_silver
            #Column cleanup
            sheet.cell(row=(coord-1+i),column=22).font, sheet.cell(row=(coord+i),column=22).fill = no_font, no_fill
        #Create comments box
        for c in range(24,33):
            for r in range(0,8):
                sheet.cell(row=coord+r,column=c).font = no_font
                sheet.cell(row=coord+r,column=c).fill = box_gray
                if sheet.cell(row=coord+r,column=c).value == "Comments (internal use):":
                    sheet.cell(row=coord+r,column=c).font = crim_note  #Ensures our original comment doesnt get reset
        start_node = sheet.cell(row=coord,column=24).coordinate
        end_node = sheet.cell(row=coord+7,column=32).coordinate
        set_border(sheet,start_node,end_node)

        for i in range(0,11):
            sheet.cell(row=coord-1,column=22+i).font = no_font
            sheet.cell(row=coord-1,column=22+i).fill = no_fill
            sheet.cell(row=coord-1,column=22+i).border = no_borders
            sheet.cell(row=coord+10,column=22+i).fill = no_fill
            sheet.cell(row=coord+10,column=22+i).font = no_font
            sheet.cell(row=coord+10,column=22+i).border = no_borders
        #bordering whole sheet
        start_node2 = sheet.cell(row=coord-1,column=1).coordinate
        end_node2 = sheet.cell(row=coord+12,column=21).coordinate
        set_border(sheet, start_node2, end_node2, True)

        sheet["B4"].value = "Standard Schedule"
        sheet["B4"].font = bold_delta
        sheet["E4"].value = "Mon:"
        sheet["E4"].font = bold_delta
        sheet["F4"].value = f'{strfdelta(mon_b_t, "{hours:02d}:{minutes:02d}")}-{strfdelta(mon_e_t, "{hours:02d}:{minutes:02d}")}'
        sheet["H4"].value = "Tue-Fri:"
        sheet["H4"].font = bold_delta
        sheet["I4"].value = f'{strfdelta(t2f_b_t, "{hours:02d}:{minutes:02d}")}-{strfdelta(t2f_e_t, "{hours:02d}:{minutes:02d}")}'
        sheet["L4"].value = "Sat:"
        sheet["L4"].font = bold_delta
        sheet["M4"].value = f'{strfdelta(sat_b_t, "{hours:02d}:{minutes:02d}")}-{strfdelta(sat_e_t, "{hours:02d}:{minutes:02d}")}'
        
    #print settings
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = True
    #inserts spacers on top of each form section; does not affect data.
    # sheet.row_dimensions.group(coord+10, coord+10, hidden=True)

def generate_summary(lst, template="Summary Template.xlsx"):                #dic format->      no. : [summary row spot, name, Rgr coord, OT coord]
    arr = dict()
    it_row = 7      #On summary template, the blank entries start at 7, this counter is used to assign "spot" for each person.
    
    wb2 = openpyxl.load_workbook(template)
    tmp = wb2.active
    mr = tmp.max_row
    mc = tmp.max_column

    for r in [sheet[coord].row for coord in lst]:
        if sheet.cell(row=r,column=12).value != None:         #Check to make sure there's an employee entry
            if sheet.cell(row=r,column=4).value == None:          #If no number assigned (manual employee), assign a number.
                sheet.cell(row=r,column=4).value = extras.pop(0)
            c = sheet.cell(row=r,column=4).value
            name = sheet.cell(row=r,column=12).value
            rgr_xy = sheet.cell(row=r,column=15).coordinate
            ot_xy =sheet.cell(row=r,column=17).coordinate
            arr[c] = [it_row, name, rgr_xy, ot_xy]
            it_row += 1
    
    period = wb["Summary"]["B2"].value
    del wb["Summary"]
    sheet2 = wb.create_sheet("Summary")
    sheet2.title = "Summary"

    for i in range(1, mr+1):
        for j in range(1, mc+1):
            j_char = openpyxl.utils.get_column_letter(j)
            sheet2.column_dimensions[j_char].width = 13
            sheet2.row_dimensions[i].height = 30
            c = tmp.cell(row = i, column = j)
            sheet2.cell(row = i, column = j).value = c.value
            sheet2.cell(row = i, column = j).font = copy(c.font)
            sheet2.cell(row = i, column = j).alignment = copy(c.alignment)
            sheet2.cell(row = i, column = j).fill = copy(c.fill)
            sheet2.cell(row = i, column = j).border = copy(c.border)

            if j in {4, 6, 8, 10}: 
                sheet2.cell(row = i, column = j).number_format = copy(c.number_format)

    for k, v in arr.items():
        sheet2.cell(row = v[0], column = 1).value = k
        sheet2.cell(row = v[0], column = 2).value = v[1]
        if t_table.get(k):
            sheet2.cell(row = v[0], column = 3).value = t_table[k][0]
        sheet2.cell(row = v[0], column = 4).value = f"='{ProcessedName}'!{v[2]}"
        sheet2.cell(row = v[0], column = 6).value = f"='{ProcessedName}'!{v[3]}"
    
    sheet2['A6'].value = f'{location.upper()}'
    
    sheet2.merge_cells('A1:O1')
    sheet2.merge_cells('A2:O2')
    sheet2.merge_cells('A3:B3')
    sheet2.merge_cells('C3:K3')
    sheet2.merge_cells('A4:A5')
    sheet2.merge_cells('B4:B5')
    sheet2.merge_cells('C4:C5')
    sheet2.merge_cells('A6:O6')
        
#Steps 
t_table = get_t_table(location)
header()
beg_markers_calc(beg_markers)
sub_header(beg_markers)
space_prep(beg_markers)
beg_markers_calc(beg_markers)
row_headers(beg_markers)
print("row_headers success")
calendar = dating(beg_markers)
print("dating success")
splitter(beg_markers)
print("splitters success")
ot_calc(beg_markers)
post_format(beg_markers)
resize()
generate_summary(beg_markers)
print("Success!")



wb.save('Output.xlsx')

#td highlight, time diff, lunch calc
